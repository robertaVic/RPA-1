# from selenium import webdriver
# from selenium.webdriver import Chrome
# from gerenciadorPlanilhas import tramitar_para_pago
import os
import shutil
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import date
from openpyxl import load_workbook
from time import sleep
import gerenciadorPastas
from os.path import isfile, join
# from gerenciadorPlanilhas import ler_dados_da_planilha, atualizar_status_na_planilha, preencher_solicitacao_na_planilha

data_em_texto = date.today().strftime("%d/%m/%Y")
arquivo_excel = gerenciadorPastas.recuperar_diretorio_usuario() +"\\tpfe.com.br\\SGP e SGC - RPA\\Resultados\\Planilha de Acompanhamento de Solicitações Financeiras 2021R.xlsx"
wb = load_workbook(arquivo_excel) #carregar o arquivo
sh1 = wb.worksheets[0]

def dados_para_tramitar_para_baixadas(tipo_de_solicitacao):
    ultima_linha = sh1.max_row
    dados = []
    # listaId = []
    todos = list(range(8, ultima_linha+1))
    for i in todos:
        tipo = sh1[f'A{i}'].value
        # todosOsIds = sh1[f"B{i}"].value
        statusRobo = sh1[f"R{i}"].value
        statusFinanceiro = sh1[f"Y{i}"].value
        # avulso = tipo[i]
        if tipo == tipo_de_solicitacao:
            # listaId.append(todosOsIds)
            statusRo = statusRobo
            statusFinan = statusFinanceiro
            if statusRo != statusFinan and statusFinan != None and statusFinan.lower() == "baixado":
                cadaSolicitacao = []
                status = i
                cadaSolicitacao.append(sh1[f"B{status}"].value) #ID
                cadaSolicitacao.append(sh1[f"D{status}"].value) #RAZAO
                cadaSolicitacao.append((sh1[f"S{status}"].value).strftime("%d/%m/%Y")) #DATA DA CRIAÇAO DE PASTA
                cadaSolicitacao.append(sh1[f"Y{status}"].value) #STATUS FINANCEIRO
                cadaSolicitacao.append(status)
                dados.append(cadaSolicitacao)
    return dados     

print(dados_para_tramitar_para_baixadas("AD")) 
# os.remove(gerenciadorPastas.recuperar_diretorio_usuario() + "\\tpfe.com.br\\SGP e SGC - RPA\\.849C9593-D756-4E56-8D6E-42412F2A707B")

# while True:
#     arquivos = gerenciadorPastas.listar_arquivos_em_diretorios(gerenciadorPastas.recuperar_diretorio_usuario() + "\\tpfe.com.br\\SGP e SGC - RPA\\")
#     if len(arquivos) > 0 and len(arquivos) != 1:
#         arquivo = arquivos[-1]
#         print(arquivo)
#         extensao = arquivo.split(".")[1]
#         if extensao == "crdownload":
#             sleep(3)
#         else:
#             try:
#                 shutil.move(gerenciadorPastas.recuperar_diretorio_usuario() + "\\tpfe.com.br\\SGP e SGC - RPA\\" + arquivo, caminho_da_pasta + data_em_texto +"\\"+ nome_da_pasta + "\\" + arquivo)
#             except:
#                 print("Não moveu o arquivo 1!")