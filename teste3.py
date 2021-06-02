import gerenciadorPastas
from openpyxl import load_workbook

arquivo_excel = gerenciadorPastas.recuperar_diretorio_usuario() +"\\tpfe.com.br\\SGP e SGC - RPA\\Configuração\\Agendamento de horarios.xlsx"


def retornar_planilha_de_log():
    wb = load_workbook(arquivo_excel) #carregar o arquivo
    #sh1 = wb.worksheets[0] #carregar a primeira planilha
    return wb

def cronograma_de_execucao():
    wb = retornar_planilha_de_log()
    sh1 = wb.worksheets[0] #carregar a primeira planilha
    #ultima_linha = sh1.max_row
    checagem1 = sh1[f'C{2}'].value
    checagem2 = sh1[f'C{3}'].value

    if checagem1 == None:
        sh1[f'C{2}'].value = "x"
    elif checagem1 != None and  checagem2 == None:
        sh1[f'C{3}'].value = "x"
    else:
        sh1[f'C{2}'].value = None
        sh1[f'C{3}'].value = None

    checagem1 = sh1[f'C{2}'].value
    checagem2 = sh1[f'C{3}'].value

    wb.save(arquivo_excel)
    print(checagem1, checagem2)

def execucao_bem_sucessida(condicao):
    cronograma_de_execucao()
    wb = retornar_planilha_de_log()
    sh1 = wb.worksheets[0]

    log1 = sh1[f'B{2}'].value
    log2 = sh1[f'B{3}'].value

    checagem1 = sh1[f'C{2}'].value
    checagem2 = sh1[f'C{3}'].value

    if log1 == None and checagem1 == "x" and condicao:
        sh1[f'B{2}'].value = "x"
        sh1[f'B{3}'].value = "x"
    elif log1 == None and checagem2 == "x" and condicao:
        sh1[f'B{2}'].value = "x"
        sh1[f'B{3}'].value = "x"

    if checagem1 == "x" and checagem2 == "x":
        sh1[f'C{2}'].value = None
        sh1[f'C{3}'].value = None
        sh1[f'B{2}'].value = None
        sh1[f'B{3}'].value = None
        wb.save(arquivo_excel)
    
    wb.save(arquivo_excel)



execucao_bem_sucessida(False)
execucao_bem_sucessida(False)
