
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.borders import Side
from gerenciadorPastas import *
from funcoes import *
from time import sleep
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

arquivo_excel = recuperar_diretorio_usuario() +"\\tpfe.com.br\\SGP e SGC - RPA\\Resultados\\Planilha de Acompanhamento de Solicitações Financeiras 2021R.xlsx"
wb = load_workbook(arquivo_excel) #carregar o arquivo
sh1 = wb.worksheets[0] #carregar a primeira planilha


def ler_dados_da_planilha(tipo_de_solicitacao):
    ultima_linha = sh1.max_row
    #print(ultima_linha1)
    #ultima_linha = sh1.cell(row=1, column=27).value
    #print(">>", ultima_linha)
    dados = []
    listaId = []
    todos = list(range(8, ultima_linha+1))
    for i in todos:
        tipo = sh1[f'A{i}'].value
        todosOsIds = sh1[f"B{i}"].value
        statusRobo = sh1[f"R{i}"].value
        statusFinanceiro = sh1[f"Y{i}"].value
        # avulso = tipo[i]
        if tipo == tipo_de_solicitacao:
            listaId.append(todosOsIds)
            statusRo = statusRobo
            statusFinan = statusFinanceiro
            if statusRo != statusFinan and statusFinan != None:
                cadaSolicitacao = []
                status = i
                #geral = [todosOsIds[status].value, sh1[f"L{status}"].value, (sh1[f"Y{status}"].value).strftime("%d/%m/%Y"), sh1[f"X{status}"].value, status]
                if (sh1[f"L{status}"].value != None) and (sh1[f"Z{status}"].value != None) and (sh1[f"Y{status}"].value != None):
                    cadaSolicitacao.append(sh1[f"B{status}"].value)
                    cadaSolicitacao.append(sh1[f"L{status}"].value)
                    cadaSolicitacao.append((sh1[f"Z{status}"].value).strftime("%d/%m/%Y"))
                    cadaSolicitacao.append(sh1[f"Y{status}"].value)
                    cadaSolicitacao.append(status)
                    dados.append(cadaSolicitacao)
    return dados     

#PARA O FLUXO DE PRESTAÇAO REALIZADA -> BAIXADAS    
def dados_para_tramitar_para_baixadas(tipo_de_solicitacao):
    ultima_linha = sh1.max_row
    dados = []
    # listaId = []
    todos = list(range(8, ultima_linha+1))
    for i in todos:
        tipo = sh1[f'A{i}'].value
        # todosOsIds = sh1[f"B{i}"].value
        statusRobo = sh1[f"R{i}"].value
        statusFinanceiro = sh1[f"Y{i}"].value
        # avulso = tipo[i]
        if tipo == tipo_de_solicitacao:
            # listaId.append(todosOsIds)
            statusRo = statusRobo
            statusFinan = statusFinanceiro
            if statusRo != statusFinan and statusFinan != None and statusFinan.lower() == "baixado":
                cadaSolicitacao = []
                status = i
                cadaSolicitacao.append(sh1[f"B{status}"].value) #ID
                cadaSolicitacao.append(sh1[f"D{status}"].value) #RAZAO
                cadaSolicitacao.append((sh1[f"S{status}"].value).strftime("%d/%m/%Y")) #DATA DA CRIAÇAO DE PASTA
                cadaSolicitacao.append(sh1[f"Y{status}"].value) #STATUS FINANCEIRO
                cadaSolicitacao.append(status)
                dados.append(cadaSolicitacao)
    return dados   

def preencher_solicitacao_na_planilha(dados_formulario, tipo_de_solicitacao):
    ultima_linha = sh1.max_row
    todos = list(range(8, ultima_linha))
    listaId = []
    listaLinha = []
    for i in todos:
        todosOsIds = sh1["B"]
        tipo = sh1['A']
        # avulso = tipo[i]
        if tipo[i].value == tipo_de_solicitacao:
            listaId.append(todosOsIds[i].value)
            listaLinha.append(todosOsIds[i].row)

    for coluna in range(len(dados_formulario)):
        #print(f"ID: {listaId[idd]} LINHA: {listaLinha[idd]}")
        if dados_formulario[1] in listaId:
            print("JA EXISTE, SOBRESCREVER")
            #sobrescrever
            linhaa = listaId.index(dados_formulario[1])
            if coluna == 12 or coluna == 13 or coluna == 14 or coluna == 18:
                if dados_formulario[coluna] == "":
                    sh1.cell(row=listaLinha[linhaa], column=coluna+1, value=dados_formulario[coluna])
                else:
                    data = datetime.strptime(dados_formulario[coluna], "%d/%m/%Y")
                    sh1.cell(row=listaLinha[linhaa], column=coluna+1, value=data)
            else:
                sh1.cell(row=listaLinha[linhaa], column=coluna+1, value=dados_formulario[coluna])
            wb.save(arquivo_excel)
        else:
            if coluna == 0:
                print("salvar novo")
            #adicionar um novo
            if coluna == 12 or coluna == 13 or coluna == 14 or coluna == 18:
                if dados_formulario[coluna] == "":
                    sh1.cell(row=ultima_linha+1, column=coluna+1, value=dados_formulario[coluna])
                else:
                    data = datetime.strptime(dados_formulario[coluna], "%d/%m/%Y")
                    sh1.cell(row=ultima_linha+1, column=coluna+1, value=data)
            else:
                sh1.cell(row=ultima_linha+1, column=coluna+1, value=dados_formulario[coluna])
            #sh1.cell(row=ultima_linha+1, column=coluna+1, value=dados_formulario[coluna])
            # print("adicionando um novo")
            wb.save(arquivo_excel)
    print("SALVOU")  
    print(ultima_linha)

def atualizar_status_na_planilha(linha):
    #pegar o status daqui mesmo
    myFill = PatternFill(start_color='A9D08E', 
                    end_color='A9D08E', 
                    fill_type = 'solid')
    status = sh1[f"Y{linha}"].value
    status = str(status).upper()
    #Status Robo
    sh1.cell(row=linha, column=18, value= status)
    #Status Financeiro
    sh1.cell(row=linha, column=25, value= status)
    #sh1.cell(row=linha, column=18, value= sh1[f"Y{linha}"].value).fill = myFill
    wb.save(arquivo_excel)        


def formatar_planilha():
    azul = PatternFill(start_color='8EA9DB', 
                    end_color='8EA9DB', 
                    fill_type = 'solid')
    
    laranja = PatternFill(start_color='F4B084', 
                    end_color='F4B084', 
                    fill_type = 'solid')
                    
    bordas = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    for celula in range(26):
        if celula +1 == 12 or celula +1 == 17 or celula +1 > 19:
            sh1.cell(row=7, column=celula+1).fill = laranja
        else:
            sh1.cell(row=7, column=celula+1).fill = azul
        sh1.cell(row=7, column=celula+1).border = bordas
    wb.save(arquivo_excel)
# def quantidade_para_tramitacao():
#     return len(ler_dados_da_planilha())        

# def tramitar_para_pago(tipo_de_solicitacao, dado, iteracao): #atualizar status
#     ler_dados_da_planilha(tipo_de_solicitacao)
#     if dado == "id":
#         return (dados[iteracao])[0]
#     elif dado == "valor":
#         return (dados[iteracao])[1]
#     elif dado == "data":
#         return (dados[iteracao])[2]     
#     elif dado == "opcao":       
#         if (dados[iteracao])[3] == "PAGO":     
#             sh1.cell(row=listaStatusDiferente[iteracao], column=18, value="PAGO")
#             wb.save(arquivo_excel)
#             return "1"
#         elif (dados[iteracao])[3] == "PARCIALMENTE PAGO":
#             sh1.cell(row=listaStatusDiferente[iteracao], column=18, value="PARCIALMENTE PAGO")
#             wb.save(arquivo_excel) 
#             return "2"





